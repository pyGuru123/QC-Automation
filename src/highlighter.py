import os
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill
from autologger import get_filename

fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

parent_dir = os.path.dirname(os.path.abspath(__file__)).strip('src\\')
test_file = os.path.join(parent_dir, 'setup.txt')

with open(test_file, 'r') as f:
    data_file = Path(f.readline())

filename = os.path.join(parent_dir, get_filename(data_file))
workbook = openpyxl.load_workbook(data_file)
sheets = workbook.sheetnames
worksheet = workbook[sheets[1]]

######################### Sorting Log File ############################
f = open(filename)
lines = f.readlines()
f.close()

with open(filename, 'w') as file:
    lines = sorted(lines, key=lambda string: int(string.split(',')[0]))
    file.writelines(lines)

####################### Highlighting Excel Rows ########################
with open(filename) as file:
    lines = file.readlines()
    for line in lines:
        row = int(line.split(',')[0])
        for cell in worksheet[row]:
            cell.fill = fill
workbook.save(data_file)

print('done')